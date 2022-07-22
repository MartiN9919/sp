import json
from copy import copy
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from converter.base.base_object import BaseTable, BaseRelation
from converter.base.base_parser import BaseParser
from converter.saphir.saphir_parser import SaphirParser
from objects.record.add_record_vector import add_data_vector, duplicates_reports
from objects.relations.add_rel import add_rel
from vector.parser_bank import ParserBank


class ConverterRelations:
    database_id_1: int # идентификатор первой базы данных (таблицы)
    database_id_2: int # идентификатор второй базы данных (таблицы)
    object_id_1: int # идентификатор первого объекта
    object_id_2: int # идентификатор второго объекта
    relations_convert_table: Dict # таблица преобразования связей (не используется)

    def __init__(self, database_id_1: int, database_id_2: int, objects_convert_table: dict,
                 relations_convert_table: dict):
        """
        Конструктор с параметрами
        @param database_id_1: идентификатор первой базы данных (таблицы)
        @param database_id_2: идентификатор второй базы данных (таблицы)
        @param objects_convert_table: таблица преобразования объектов (старые идентификаторы в новые)
        @param relations_convert_table: таблица преобразования связей (не используется)
        """
        self.database_id_1 = database_id_1
        self.database_id_2 = database_id_2
        self.object_id_1 = objects_convert_table.get(database_id_1)
        self.object_id_2 = objects_convert_table.get(database_id_2)
        self.relations_convert_table = relations_convert_table


    def convert(self, relation: BaseRelation, objects: Dict, key_id: int, value='') -> dict:
        """
        Метод для конвертирования связи Хроноса в связи Сапфира
        @param relation: связи Хроноса
        @param objects: словарь соответствия объектов
        @param key_id: идентификатор связи
        @param value: значение связи
        @return: словарь с информацией о созданной связи
        """
        if self.object_id_1 is None or self.object_id_2 is None:
            return {}
        old_id_1 = relation.rec_id_1
        old_id_2 = relation.rec_id_2
        new_object_1 = objects.get(f"{self.database_id_1}_{old_id_1}")
        new_object_2 = objects.get(f"{self.database_id_2}_{old_id_2}")
        if new_object_1 and new_object_2:
            new_id_1 = new_object_1['rec_id']
            new_id_2 = new_object_2['rec_id']
            date_time = relation.date + ' ' + relation.time
            try:
                add_rel(1, self.object_id_1, new_id_1, self.object_id_2, new_id_2, [{'id': key_id, 'value': value, 'date': date_time}])
            except Exception as e:
                print(f"error creating relation {self.database_id_1}_{old_id_1}_{self.database_id_2}_{old_id_2}")
                return {
                    'error': {
                        f"{self.database_id_1}_{old_id_1}_{self.database_id_2}_{old_id_2}": {
                            'object_id_1': self.object_id_1,
                            'rec_id_1': new_id_1,
                            'object_id_2': self.object_id_2,
                            'rec_id_2': new_id_2,
                            'key_id': key_id,
                            'date': date_time
                        }
                    },
                }
            result: Dict = {
                'object_id_1': self.object_id_1,
                'rec_id_1': new_id_1,
                'object_id_2': self.object_id_2,
                'rec_id_2': new_id_2,
                'key_id': key_id,
                'date': date_time
            }
            return result
        else:
            return {}

    @property
    def id(self):
        return f"{self.database_id_1}_{self.database_id_2}"


class ConverterParams:
    database_id: int # идентификатор базы данных (таблицы)
    object_id: int # идентификатор объекта Сапфир
    params_converter_table: Dict[int, Tuple[
        int, int]]  # таблица соответствия старого id и нового id с возможным порядком следования (например ФИО)

    def __init__(self, database_id: int, object_id: int, params_convert_dict: dict):
        """
        Конструктор с параметрами
        @param database_id: идентификатор базы данных (таблицы)
        @param object_id: идентификатор объекта Сапфир
        @param params_convert_dict: таблица соответствия старого id и нового id с возможным порядком следования (например ФИО)
        """
        self.database_id = database_id
        self.object_id = object_id
        self.params_converter_table = params_convert_dict

    def convert(self, database: BaseTable, path: str) -> Dict:
        """
        Метод для преобразования параметров в записи Сапфира
        @param database: база данных (таблица)
        @param path: путь к папке с файлами
        @return: словарь с информацией о созданных объектах и отложенных связях
        """
        result: Dict[str, Dict] = {'objects': {}, 'relations': {}, 'error': {}}
        if self.object_id == 0 or len(self.params_converter_table) == 0:
            return result
        for database_object_id in database.data:
            key = f"{self.database_id}_{database_object_id}"
            new_object = {key: {'old_id': key, 'rec_id': 0, 'object_id': self.object_id, 'params': []}}
            database_object_params = database.data[database_object_id]['values']
            database_object_datetime = database.data[database_object_id].get('date', '01-01-1900') + ' ' + \
                                       database.data[database_object_id].get('time', '00:00')
            temp_params = {}
            for param in database_object_params:
                new_param_info = self.params_converter_table.get(param['type'].id, (0, 0))
                if new_param_info[0] == 0 and param['type'].param_type == 'Ф':
                    new_param_info = (1, 0)
                elif new_param_info[0] == 0:
                    param['value'] = param['type'].name + ': ' + param['value']
                if new_param_info[1]:
                    if not temp_params.get(new_param_info[0]):
                        temp_params[new_param_info[0]] = [(param['value'], new_param_info[1])]
                    else:
                        temp_params[new_param_info[0]].append((param['value'], new_param_info[1]))
                else:
                    new_object[key]['params'].append(
                        {'id': new_param_info[0], 'value': param['value'],
                         'date': param.get('datetime', database_object_datetime)})
            for temp_param in temp_params:
                value = ' '.join([item[0] for item in sorted(temp_params[temp_param], key=lambda x: x[1])])
                new_object[key]['params'].append(
                    {'id': temp_param, 'value': value, 'date': database_object_datetime})
            if len(new_object[key]['params']) == 0: # toDo перенести в парсер
                continue
            temp_result = add_data_vector(1, new_object[key], path)
            if temp_result.get('object'):
                new_object[key]['rec_id'] = temp_result['object']
            else:
                print('some error ', key, ' ', new_object[key])
                result['error'][key] = new_object[key]
            result['objects'].update({key: {'rec_id': new_object[key]['rec_id']}})
        return result


class ConverterBank:
    converter_type: str
    bank_parser: BaseParser
    converter_objects_table: Dict[int, int] = {}  # словарь соответствия объектов
    converter_relations_table: Dict = {}  # словарь соответствия связей ('о1_о2') -> (key_id)
    converters_params: Dict[int, ConverterParams] = {}  # словарь соответсия параметров объекта (o.id) -> ConverterParams
    converters_relations: Dict[str, ConverterRelations] = {}
    object_to_create: Dict = {} # накопитель результата созданных объектов
    relation_to_create: List = [] # накопитель результата созданных связей

    def __init__(self, convert_setting: dict):
        """
        Конструктор с параметром
        @param convert_setting: словарь содержащий исходные параметры для конвертора и парсера
        """
        self.converter_type = convert_setting.get('converter_type', 'chronos')
        if convert_setting.get('init_type', 'exel') == 'exel': # если тип таблицы преобразования exel
            self._parse_exel(convert_setting['convert_path'])
        else: # если тип таблицы преобразования не поддерживается
            raise TypeError
        if self.converter_type == 'chronos':
            self.bank_parser = ParserBank(convert_setting['parser'])
        elif self.converter_type == 'saphir':
            self.bank_parser = SaphirParser(convert_setting['parser'])

    def _parse_exel(self, path: str):
        """
        Метод для парсинга таблицы преобразования формата exel
        @param path: путь к exel файлу
        """
        work_book = load_workbook(path)
        objects = work_book['objects']
        params_convert_dict, row0, row1 = {}, 0, 0
        for row in objects.iter_rows():
            if isinstance(row[0].value, int):
                self.converter_objects_table[row[0].value] = row[1].value
                if row0:
                    self.converters_params[row0] = ConverterParams(row0, row1, copy(params_convert_dict))
                    params_convert_dict = {}
                row0, row1 = row[0].value, row[1].value
            elif isinstance(row[2].value, int):
                params_convert_dict[row[2].value] = (row[3].value, row[4].value)
        else:
            self.converters_params[row0] = ConverterParams(row0, row1, copy(params_convert_dict))
        relations = work_book['relations']
        if self.converter_type == 'chronos':
            for row in relations.iter_rows():
                if isinstance(row[0].value, int):
                    data_base_id_1 = row[0].value
                    data_base_id_2 = row[1].value
                    key_id = row[2].value
                    if data_base_id_2 < data_base_id_1:
                        data_base_id_1, data_base_id_2 = data_base_id_2, data_base_id_1
                    self.converter_relations_table[f"{data_base_id_1}_{data_base_id_2}"] = key_id
        elif self.converter_type == 'saphir':
            key_id_1, key_id_2, values = 0, 0, {}
            for row in relations.iter_rows():
                if isinstance(row[0].value, int):
                    if key_id_1 != 0:
                        self.converter_relations_table[key_id_1] = {'key_id': key_id_2, 'values': values}
                    key_id_1 = row[0].value
                    key_id_2 = row[2].value
                    values = {}
                if isinstance(row[1].value, int):
                    values[row[1].value] = row[3].value
            else:
                self.converter_relations_table[key_id_1] = {'key_id': key_id_2, 'values': values}


    def convert(self):
        """
        Метод для преобразования данных/связей Хроноса в Сапфир
        """
        errors: dict = {}
        for database in self.bank_parser.databases:
            if self.converters_params.get(database.id) is None:
                continue
            converter_params = self.converters_params[database.id]
            temp = converter_params.convert(database, self.bank_parser.documents_path)
            self.object_to_create.update(temp['objects'])
            errors.update(temp['error'])
        for relation in self.bank_parser.relations.values():
            converter_relation = self.converters_relations.get(f"{relation.object_id_1}_{relation.object_id_2}")
            if not converter_relation:
                converter_relation = ConverterRelations(relation.object_id_1, relation.object_id_2,
                                                        self.converter_objects_table, self.converter_relations_table)
                self.converters_relations[f"{relation.object_id_1}_{relation.object_id_2}"] = converter_relation
            key_id = 0
            value = ''
            if self.converter_type == 'chronos':
                key_id = self.converter_relations_table.get(f"{relation.object_id_1}_{relation.object_id_2}", 0)
            elif self.converter_type == 'saphir':
                relation_info = self.converter_relations_table.get(relation.key_id, {'key_id': 0, 'values': {}})
                key_id = relation_info['key_id']
                value = relation_info['values'].get(relation.value, 0)
            temp = converter_relation.convert(relation, self.object_to_create, key_id, value)
            if temp.get('error'):
                errors.update(temp['error'])
            elif len(temp) > 0:
                self.relation_to_create.append(temp)
        with open('/deploy_storage/object_convert.json', 'w') as file:
            json.dump(self.object_to_create, file)


CONVERT_SETTING = json.loads(open('param_saphir.json').read().encode().decode('utf-8-sig'))
converter = ConverterBank(CONVERT_SETTING)
print('start_converting')
# converter.convert()
#
# report_file = open("/deploy_storage/report.txt", "w")
# for report in duplicates_reports:
#     report_file.write(report + '\n')
# report_file.close()

