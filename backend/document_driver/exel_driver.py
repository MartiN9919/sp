import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from core.projectSettings.constant import DOCUMENT_ROOT, TEMPLATE_ROOT


def insert_list(works_sheet, mode, start_cell, data):
    """
    Функция для вставки списка в exel документ
    @param works_sheet: рабочее окно для вставки
    @param mode: режим вставки (ud, down, left, right)
    @param start_cell: стартовая ячейка
    @param data: список для вставки
    """
    col = start_cell.col_idx
    row = start_cell.row
    if mode == 'down' or mode == 'up':
        for index, value in enumerate(data):
            index = index if mode == 'down' else -index
            works_sheet.cell(row + index, col).value = value
    if mode == 'left' or mode == 'right':
        for index, value in enumerate(data):
            index = index if mode == 'right' else -index
            works_sheet.cell(row, col + index - 1).value = value


def as_text(value):
    if value is None:
        return ""
    return str(value)


def reorder_work_sheet(work_sheet):
    """
    Функция для установки ширины столбцов по максимальной ширине содержимого
    @param work_sheet: страница на которой необходимо установить ширину
    """
    for column_cells in work_sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells) + 1
        work_sheet.column_dimensions[column_cells[0].column_letter].width = length


def merge_column_by_same_value(index, work_sheet):
    """
    Функция для объединения ячеек с одинаковым содержимым в заданном столбце
    @param index: индекс столбца
    @param work_sheet: страница документа на которой требуется объекдинение
    """
    temp = list(work_sheet.rows)[1][index]
    start_value = temp.value
    start_cell = temp.coordinate
    prev_cell = temp.coordinate
    for row in list(work_sheet.rows)[2:]:
        if row[index].value != start_value:
            work_sheet.merge_cells(start_cell + ':' + prev_cell)
            work_sheet[start_cell].alignment = Alignment(vertical='top')
            start_cell = row[index].coordinate
            prev_cell = row[index].coordinate
            start_value = row[index].value
        else:
            prev_cell = row[index].coordinate
    work_sheet.merge_cells(start_cell + ':' + prev_cell)
    work_sheet[start_cell].alignment = Alignment(vertical='top')


def get_xlsx_document_from_template(template_path, name, data):
    """
    Функция для формирования exel документа из шаблона
    @param template_path: название шаблона
    @param name: название конечного документа
    @param data: словарь содержащий информаию для занесение в документ, формат:
    {Position:[data1,...,dataN],...,PositionN:p[]} , где Position позиция в шаблоне, data -  данные для занесения
    @return: путь к конечному документу
    """
    work_book = load_workbook(TEMPLATE_ROOT + template_path)
    works_sheet = work_book.active
    for col in works_sheet.iter_cols():
        for row in col:
            if row.value and re.sub("<[^<]+>", "", str(row.value)) in data.keys():
                mode = row.value[row.value.find('<') + 1:row.value.find('>')] if row.value.find('<') != -1 else None
                if not mode:
                    row.value = data[row.value][0]
                else:
                    insert_list(works_sheet, mode, row, data[re.sub("<[^<]+>", "", row.value)])
    reorder_work_sheet(works_sheet)
    work_book.save(DOCUMENT_ROOT + name + '.xlsx')
    return DOCUMENT_ROOT + name + '.xlsx'





