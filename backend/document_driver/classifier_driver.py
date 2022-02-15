from openpyxl import Workbook

from core.settings import DOCUMENT_ROOT
from data_base_driver.sys_key.get_list import get_lists
from data_base_driver.sys_key.get_object_info import obj_list
from document_driver.exel_driver import reorder_work_sheet, merge_column_by_same_value
from objects.record.get_record import get_keys_blank
from objects.relations.get_rel import get_relations_list


def make_objects_sheet(sheet):
    """
    Функция для формирования листа объектов
    @param sheet: пустой лист
    """
    objects = obj_list()
    sheet.append(['id', 'title_single', 'title'])
    for row in [[item['id'], item['title_single'], item['title']] for item in objects]:
        sheet.append(row)
    reorder_work_sheet(sheet)


def make_prop_sheet(sheet):
    """
    Функция для формирования листа классификаторов
    @param sheet: пустой лист
    """
    classifiers = [item for item in sorted(get_keys_blank(), key=lambda x: x['obj_id']) if not item['blocked_blank']]
    sheet.append(['obj_id', 'id', 'title', 'hint', 'type', 'list_id'])
    for row in [[item['obj_id'], item['id'], item['title'], item['hint'], item['type'], item['list_id']]
                for item in classifiers]:
        sheet.append(row)
    merge_column_by_same_value(0, sheet)
    reorder_work_sheet(sheet)


def make_relation_sheet(sheet):
    """
    Функция для формирования листа связей
    @param sheet: пустой лист
    """
    objects = obj_list()
    objects.append({'id': 1})  # костыль, так как в списке объектов нет связи
    relations = [item for item in get_relations_list() if not item['blocked_blank']]
    all_relations = []
    for relation in relations:
        if relation['object_id_1'] != relation['object_id_2']:
            all_relations.append({
                'object_id_1': relation['object_id_2'],
                'object_id_2': relation['object_id_1'],
                'id': relation['id'],
                'title': relation['title'],
                'hint': relation['hint'],
                'type': relation['type'],
            })
    all_relations += relations
    all_relations = sorted(sorted(all_relations, key=lambda x: x['object_id_2']), key=lambda x: x['object_id_1'])
    sheet.append(['object_id_1', 'object_id_2', 'id', 'title', 'hint', 'type', 'list_id'])
    for row in [[item['object_id_1'],
                 item['object_id_2'],
                 item['id'],
                 item['title'],
                 item['hint'],
                 item['type']['title'] if item['type']['title'] != 'unknow' else None,
                 item['type']['value'],
                 ] for item in all_relations]:
        sheet.append(row)
    merge_column_by_same_value(0, sheet)
    merge_column_by_same_value(1, sheet)
    reorder_work_sheet(sheet)


def make_list_sheet(sheet):
    """
    Функция для формирования листа списков
    @param sheet: пустой лист
    """
    lists = get_lists()
    sheet.append(['id', 'name', 'title', 'val_id', 'value'])
    for list_id in lists:
        list_params = lists[list_id]
        for row in ([[list_id,
                      list_params['name'],
                      list_params['title'],
                      item['id'],
                      item['value']
                      ] for item in list_params['values']]
        ):
            sheet.append(row)
    merge_column_by_same_value(0, sheet)
    merge_column_by_same_value(1, sheet)
    merge_column_by_same_value(2, sheet)
    reorder_work_sheet(sheet)


def get_exel_document(name):
    """
    Функция для формирования файла классификаторов
    @param name: имя выходного файла
    @return: путь к выходному файлу
    """
    work_book = Workbook()
    obj_sheet = work_book.create_sheet('obj')
    make_objects_sheet(obj_sheet)
    prop_sheet = work_book.create_sheet('prop')
    make_prop_sheet(prop_sheet)
    rel_sheet = work_book.create_sheet('rel')
    make_relation_sheet(rel_sheet)
    list_sheet = work_book.create_sheet('lst')
    make_list_sheet(list_sheet)
    work_book.remove(work_book['Sheet'])
    work_book.save(DOCUMENT_ROOT + name + '.xlsm')
    return DOCUMENT_ROOT + name + '.xlsm'
