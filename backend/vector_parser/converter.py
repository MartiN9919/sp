import json
from copy import copy
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from data_bank import Database, Relation
from objects.record.add_record_vector import add_data_vector
from objects.relations.add_rel import add_rel
from parser_bank import ParserBank


class ConverterRelations:
    database_id_1: int # идентификатор первой базы данных (таблицы)
    database_id_2: int # идентификатор второй базы данных (таблицы)
    object_id_1: int # идентификатор первого объекта
    object_id_2: int # идентификатор второго объекта
    relations_convert_table: Dict # таблица преобразования связей (не используется)
    deferred_relations: Dict[str, Dict[str, Dict[str, list]]] # отложенные связи

    def __init__(self, database_id_1: int, database_id_2: int, objects_convert_table: dict,
                 relations_convert_table: dict, deferred_relations: dict):
        """
        Конструктор с параметрами
        @param database_id_1: идентификатор первой базы данных (таблицы)
        @param database_id_2: идентификатор второй базы данных (таблицы)
        @param objects_convert_table: таблица преобразования объектов (старые идентификаторы в новые)
        @param relations_convert_table: таблица преобразования связей (не используется)
        @param deferred_relations: отложенные связи
        """
        self.database_id_1 = database_id_1
        self.database_id_2 = database_id_2
        self.object_id_1 = objects_convert_table[database_id_1]
        self.object_id_2 = objects_convert_table[database_id_2]
        self.relations_convert_table = relations_convert_table
        self.deferred_relations = deferred_relations

    def create_deferred_relation(self, obj_1: str, obj_2: str, convert_table: dict, objects: dict, relation: Relation) -> None:
        """
        Метод для создания отложенной связи (преобразования 2-х связей в одну именованную)
        @param obj_1: идентификатор первого объекта
        @param obj_2: идентификатор второго объекта
        @param convert_table: таблица преобразования связей
        @param objects: таблица соответствия старых-новых объектов
        @param relation: преобразуемая связь
        """
        if obj_1 in convert_table:
            obj, database_id, rec_id = obj_1, self.database_id_2, relation.rec_id_2
        else:
            obj, database_id, rec_id = obj_2, self.database_id_1, relation.rec_id_1
        if not self.deferred_relations.get(obj):
            self.deferred_relations[obj] = {}
        for convert_relation in convert_table[obj]:
            key = f"{convert_relation['key_id']}_{convert_relation['value']}"
            if not self.deferred_relations[obj].get(key):
                self.deferred_relations[obj][key] = {
                    'rec_id_1': [],
                    'rec_id_2': [],
                    'date': convert_relation['date']
                }
            if convert_relation['rel_id_1'] == database_id:
                self.deferred_relations[obj][key]['rec_id_1'].append(objects[f"{database_id}_{rec_id}"]['rec_id'])
                self.deferred_relations[obj][key]['object_id_1'] = objects[f"{database_id}_{rec_id}"]['object_id']
            elif convert_relation['rel_id_2'] == database_id:
                self.deferred_relations[obj][key]['rec_id_2'].append(objects[f"{database_id}_{rec_id}"]['rec_id'])
                self.deferred_relations[obj][key]['object_id_2'] = objects[f"{database_id}_{rec_id}"]['object_id']
            if convert_relation['document'] == database_id:
                self.deferred_relations[obj][key]['document'] = objects[f"{database_id}_{rec_id}"]['rec_id']

    def convert(self, relation: Relation, objects: Dict, convert_table: dict) -> dict:
        """
        Метод для конвертирования связи Хроноса в связи Сапфира
        @param relation: связи Хроноса
        @param objects: словарь соответствия объектов
        @param convert_table: таблица формирования отложенных связей
        @return: словарь с информацией о созданной связи
        """
        obj_1 = f"{self.database_id_1}_{relation.rec_id_1}"
        obj_2 = f"{self.database_id_2}_{relation.rec_id_2}"
        if obj_1 in convert_table or obj_2 in convert_table:
            self.create_deferred_relation(obj_1, obj_2, convert_table, objects, relation)
        old_id_1 = relation.rec_id_1
        old_id_2 = relation.rec_id_2
        new_object_1 = objects.get(f"{self.database_id_1}_{old_id_1}")
        new_object_2 = objects.get(f"{self.database_id_2}_{old_id_2}")
        if new_object_1 and new_object_2:
            new_id_1 = new_object_1['rec_id']
            new_id_2 = new_object_2['rec_id']
            key_id = self.relations_convert_table.get(f"{self.database_id_1}_{self.database_id_2}", 0)
            date_time = relation.date + ' ' + relation.time

            add_rel(1, self.object_id_1, new_id_1, self.object_id_2, new_id_2, [{'id': key_id, 'value': '', 'date': date_time}])

            result: Dict = {
                'object_id_1': self.object_id_1,
                'rec_id_1': new_id_1,
                'object_id_2': self.object_id_2,
                'rec_id_2': new_id_2,
                'key_id': key_id,
                'date': date_time
            }
            return result
        else:
            return {}

    @property
    def id(self):
        return f"{self.database_id_1}_{self.database_id_2}"


class ConverterParams:
    database_id: int # идентификатор базы данных (таблицы)
    object_id: int # идентификатор объекта Сапфир
    params_converter_table: Dict[int, Tuple[
        int, int]]  # таблица соответствия старого id и нового id с возможным порядком следования (например ФИО)

    def __init__(self, database_id: int, object_id: int, params_convert_dict: dict):
        """
        Конструктор с параметрами
        @param database_id: идентификатор базы данных (таблицы)
        @param object_id: идентификатор объекта Сапфир
        @param params_convert_dict: таблица соответствия старого id и нового id с возможным порядком следования (например ФИО)
        """
        self.database_id = database_id
        self.object_id = object_id
        self.params_converter_table = params_convert_dict

    @staticmethod
    def _add_relations(database_id: int, data: dict, converters_object_to_relation: dict, result: dict):
        """
        Метод дял добавления отложенной связи
        @param database_id: идентификатор базы данных (таблицы)
        @param data: словарь содержащий объекты и их параметры
        @param converters_object_to_relation: таблица преобразования объектов в связи
        @param result: словарь для накопления результата
        """
        for database_object_id in data:
            database_object_params = data[database_object_id]['values']
            key = f"{database_id}_{database_object_id}"
            relation = {key: []}
            for relation_converter in converters_object_to_relation[database_id]:
                if not relation_converter['date']:
                    date = f"{data[database_object_id]['date']} {data[database_object_id]['time']}"
                else:
                    try:
                        date = next(item for item in database_object_params
                                    if item['type'].id == relation_converter['date'])['value'] + ' 00:00'
                    except StopIteration as e:
                        continue
                temp = {
                    'rel_id_1': relation_converter['rel_id_1'],
                    'rel_id_2': relation_converter['rel_id_2'],
                    'key_id': relation_converter['key_id'],
                    'value': relation_converter['value'],
                    'date': date,
                    'document': relation_converter['document']
                }
                relation[key].append(temp)
            result['relations'].update(relation)

    def convert(self, database: Database, converters_object_to_relation: dict, path: str) -> Dict:
        """
        Метод для преобразования параметров в записи Сапфира
        @param database: база данных (таблица)
        @param converters_object_to_relation: таблица преобразования объектов в связи
        @param path: путь к папке с файлами
        @return: словарь с информацией о созданных объектах и отложенных связях
        """
        result: Dict[str, Dict] = {'objects': {}, 'relations': {}}
        if database.id in converters_object_to_relation:
            ConverterParams._add_relations(database.id, database.data, converters_object_to_relation, result)
        if self.object_id == 0 or len(self.params_converter_table) == 0:
            return result
        for database_object_id in database.data:
            key = f"{self.database_id}_{database_object_id}"
            new_object = {key: {'rec_id': 0, 'object_id': self.object_id, 'params': []}}
            database_object_params = database.data[database_object_id]['values']
            database_object_datetime = database.data[database_object_id]['date'] + ' ' + \
                                       database.data[database_object_id]['time']
            temp_params = {}
            for param in database_object_params:
                new_param_info = self.params_converter_table.get(param['type'].id, (0, 0))
                if new_param_info[0] == 0 and param['type'].param_type == 'Ф':
                    new_param_info = (1, 0)
                elif new_param_info[0] == 0:
                    param['value'] = param['type'].name + ': ' + param['value']
                if new_param_info[1]:
                    if not temp_params.get(new_param_info[0]):
                        temp_params[new_param_info[0]] = [(param['value'], new_param_info[1])]
                    else:
                        temp_params[new_param_info[0]].append((param['value'], new_param_info[1]))
                else:
                    new_object[key]['params'].append(
                        {'id': new_param_info[0], 'value': param['value'], 'date': database_object_datetime})
            for temp_param in temp_params:
                value = ' '.join([item[0] for item in sorted(temp_params[temp_param], key=lambda x: x[1])])
                new_object[key]['params'].append(
                    {'id': temp_param, 'value': value, 'date': database_object_datetime})
            if len(new_object[key]['params']) == 0: # toDo перенести в парсер
                continue
            temp_result = add_data_vector(1, new_object[key], path)
            if temp_result.get('object'):
                new_object[key]['rec_id'] = temp_result['object']
            else:
                print('some error ', key, ' ', new_object[key])
            result['objects'].update(new_object)
        return result


class ConverterBank:
    bank_parser: ParserBank
    converter_objects_table: Dict[int, int] = {}  # словарь соответствия объектов
    converter_relations_table: Dict[str, int] = {}  # словарь соответствия связей ('о1_о2') -> (key_id)
    converters_params: Dict[int, ConverterParams] = {}  # словарь соответсия параметров объекта (o.id) -> ConverterParams
    converters_relations: Dict[str, ConverterRelations] = {}
    converters_object_to_relation: Dict[int, List[dict]] = {}
    object_to_create: Dict = {}
    relation_to_create: List = []

    def __init__(self, convert_setting: dict):
        if convert_setting.get('init_type', 'exel') == 'exel':
            self._parse_exel(convert_setting['convert_path'])
        else:
            raise TypeError
        self.bank_parser = ParserBank(convert_setting['parser'])

    def _parse_exel(self, path: str):
        work_book = load_workbook(path)
        objects = work_book['objects']
        params_convert_dict, row0, row1 = {}, 0, 0
        for row in objects.iter_rows():
            if isinstance(row[0].value, int):
                self.converter_objects_table[row[0].value] = row[1].value
                if row[2].value and row[2].value == 1:
                    if not self.converters_object_to_relation.get(row[0].value):
                        self.converters_object_to_relation[row[0].value] = []
                if row0:
                    self.converters_params[row0] = ConverterParams(row0, row1, copy(params_convert_dict))
                    params_convert_dict = {}
                row0, row1 = row[0].value, row[1].value
            elif isinstance(row[2].value, int):
                params_convert_dict[row[2].value] = (row[3].value, row[4].value)
        else:
            self.converters_params[row0] = ConverterParams(row0, row1, copy(params_convert_dict))
        if len(self.converters_object_to_relation) > 0:
            relations = work_book['relations']
            for row in relations.iter_rows():
                if isinstance(row[0].value, int):
                    data_base_id = row[0].value
                    self.converters_object_to_relation[data_base_id].append({
                        'rel_id_1': row[1].value,
                        'rel_id_2': row[2].value,
                        'key_id': row[3].value,
                        'value': row[4].value if row[4].value else '',
                        'date': row[5].value,
                        'document': row[6].value
                    })

    def create_deferred_relations(self, deferred_relations: dict):
        for deferred_relation in deferred_relations.values():
            for key in deferred_relation:
                key_list = key.split('_')
                key_id, value = int(key_list[0]), int(key_list[1])
                date = deferred_relation[key]['date']
                object_id_1 = deferred_relation[key]['object_id_1']
                object_id_2 = deferred_relation[key]['object_id_2']
                rec_id_1_list = deferred_relation[key]['rec_id_1']
                rec_id_2_list = deferred_relation[key]['rec_id_2']
                document = deferred_relation[key].get('document')
                for rec_id_1 in rec_id_1_list:
                    for rec_id_2 in rec_id_2_list:
                        add_rel(1, object_id_1, rec_id_1, object_id_2, rec_id_2,
                                [{'id': key_id, 'value': value, 'date': date}], document)
                        self.relation_to_create.append({
                            'object_id_1': object_id_1,
                            'rec_id_1': rec_id_1,
                            'object_id_2': object_id_2,
                            'rec_id_2': rec_id_2,
                            'key_id': key_id,
                            'value': value,
                            'datetime': date
                        })

    def convert(self):
        temp_relations: dict = {}
        for database in self.bank_parser.databases:
            converter_params = self.converters_params[database.id]
            temp = converter_params.convert(database, self.converters_object_to_relation,
                                            self.bank_parser.documents_path)
            self.object_to_create.update(temp['objects'])
            temp_relations.update(temp['relations'])
        deferred_relations: dict = {}
        for relation in self.bank_parser.relations.values():
            converter_relation = self.converters_relations.get(f"{relation.object_id_1}_{relation.object_id_2}")
            if not converter_relation:
                converter_relation = ConverterRelations(relation.object_id_1, relation.object_id_2,
                                                        self.converter_objects_table, self.converter_relations_table,
                                                        deferred_relations)
                self.converters_relations[f"{relation.object_id_1}_{relation.object_id_2}"] = converter_relation
            temp = converter_relation.convert(relation, self.object_to_create, temp_relations)
            if len(temp) > 0:
                self.relation_to_create.append(temp)
        self.create_deferred_relations(deferred_relations)

CONVERT_SETTING = json.load(open('/home/pushkin/converter_param.json'))
converter = ConverterBank(CONVERT_SETTING)
converter.convert()


