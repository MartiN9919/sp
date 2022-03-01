# -*- coding: utf-8 -*-

from   openpyxl               import Workbook, load_workbook
from   openpyxl.comments      import Comment
from   openpyxl.drawing.image import Image
#from   openpyxl.worksheet.protection import SheetProtection

from   fun.funSys             import lstDim2

#####################################################
# https://openpyxl.readthedocs.io/en/stable/
# pip install openpyxl [--upgrade]
# ver 2.5.2
#####################################################
# pip install pillow - для работы с картинками
# ver 5.1.0
#####################################################
# openpyxl==2.5.2
# et-xmlfile==1.0.1
# jdcal==1.3
# Pillow==5.1.0
#####################################################



#####################################################
# РАБОТА С EXCEL
#####################################################
# filePath - путь к файлу, если не задан, то создать файл
#####################################################
class Excel():
    def __init__(self, filePath=''):
        self.filePath = filePath
        self.wb = Workbook() if filePath.strip()=='' else load_workbook(filename = filePath) #, keep_vba=True  Workbook(write_only=True)

    def save(self, filePath=''):
        self.wb.save(filePath if filePath != '' else self.filePath)



    #####################################################
    # ЗАПИСАТЬ ЗНАЧЕНИЯ
    #####################################################
    # data      - значение, одно или двухмерный массив значений
    # transpose - транспонирование перед выводом
    #####################################################
    def valueWrite(self, addrFull, data, transpose=False):
        addr = self.addrSplit(addrFull)
        return self._valueWrite_(addr[0], addr[1], data, transpose)

    def _valueWrite_(self, wsName, addr, data, transpose=False):
        data = lstDim2(data)                                                    # в двухмерный список
        if transpose: data = list(zip(*data))                                   # транспонирование

        cell = self.wb[wsName][addr]
        for row, data_row in enumerate(data):
            for col, data_cell in enumerate(data_row):
                cell.offset(column=col, row=row).value = data_cell



    #####################################################
    # ЧИТАТЬ ЗНАЧЕНИЯ
    #####################################################
    # dx, dy  - размер читаемой области
    # возврат - массив значений [[],[],...] или
    #           значение, если dx, dy не заданы или = 1
    #####################################################
    def dataRead(self, addrFull, dx=1, dy=1):
        addr = self.addrSplit(addrFull)
        return self._dataRead_(addr[0], addr[1], dx, dy)

    def _dataRead_(self, wsName, addr, dx=1, dy=1):
        cell = self.wb[wsName][addr]
        if (dx!=1) or (dy!=1):
            ret = []
            for row in range(1, dy):
                ret_row = []
                for col in range(1, dx):
                    ret_row.append(cell.offset(column=col, row=row).value)
                ret.append(ret_row)
        else:
            ret = cell.value
        return ret



    #####################################################
    # ЗАПИСАТЬ КОММЕТАРИИ
    #####################################################
    # data      - значение, одно или двухмерный массив значений
    # transpose - транспонирование перед выводом
    #####################################################
    def commentWrite(self, addrFull, data, transpose=False):
        addr = self.addrSplit(addrFull)
        return self._commentWrite_(addr[0], addr[1], data, transpose)

    def _commentWrite_(self, wsName, addr, data, transpose=False):
        data = lstDim2(data)                                                    # в двухмерный список
        if transpose: data = list(zip(*data))                                   # транспонирование

        cell = self.wb[wsName][addr]
        for row, data_row in enumerate(data):
            for col, data_cell in enumerate(data_row):
                if data_cell != '': cell.offset(column=col, row=row).comment = Comment(data_cell, "")




    #####################################################
    # ЗАПИСАТЬ КАРТИНКУ
    #####################################################
    # data      - пути к картинкам: значение, одно или двухмерный массив значений
    # transpose - транспонирование перед выводом
    # maxSize   - [width, height]  значения могут = 0 - без измениения
    #####################################################
    def imageWrite(self, addrFull, data, transpose=False, maxSize = []):
        addr = self.addrSplit(addrFull)
        return self._imageWrite_(addr[0], addr[1], data, transpose, maxSize)

    def _imageWrite_(self, wsName, addr, data, transpose=False, maxSize = []):
        data = lstDim2(data)                                                    # в двухмерный список
        if transpose: data = list(zip(*data))                                   # транспонирование

        cell = self.wb[wsName][addr]
        for row, data_row in enumerate(data):
            for col, data_cell in enumerate(data_row):
                if data_cell == '': continue
                #cell.offset(column=col, row=row).comment = Comment(data_cell, "")

                try:
                    img = Image(data_cell)

                    if len(maxSize) == 2:
                        maxWidth   = maxSize[0] if maxSize[0] > 0 else img.width
                        maxHeight  = maxSize[1] if maxSize[1] > 0 else img.height
                        ratio      = min(maxWidth / float(img.width), maxHeight / float(img.height))

                        img.width  = int(img.width  * ratio)
                        img.height = int(img.height * ratio)

                    self.wb[wsName].add_image(img, cell.offset(column=col, row=row).coordinate)
                except:
                    pass                                                        # запрещаем ошибку при отсутствии файла



    #####################################################
    # РАЗБИТЬ АДРЕС НА ЛИСТ И ЯЧЕЙКИ
    #####################################################
    # addrFull = 'Лист 1'!A1  |  "Лист 1"!$A$1  |  Лист!A1
    #####################################################
    addrSplit = lambda self, addrFull: addrFull.replace("'", '').replace('"', '').split('!')


    #####################################################
    # ЗАЩИТИТЬ ЛИСТ
    #####################################################
    # НЕ РАБОТАЕТ, ПИШУТ ПРО БАГ
    #####################################################
    #def wsProtectionOn(self, wsName, password=''):
    #    protection = self.wb[wsName].protection
    #    protection = SheetProtection()
    #    protection.set_password(password)
    #    protection.enable()
    #    protection.selectLockedCells = True
    #    protection.sheet = True



    #arr = self.ws['A1':'A5']
    #wb.protection.sheet = True
